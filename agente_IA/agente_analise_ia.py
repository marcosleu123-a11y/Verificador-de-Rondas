import argparse
import base64
import csv
import json
import mimetypes
import os
import tempfile
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional


COLUNA_ANALISADA = "analisada por IA"
COLUNA_GRUPO = "grupo_analise_ia"
COLUNA_CONFIANCA = "confianca_analise_ia"
COLUNA_MOTIVO = "motivo_analise_ia"
COLUNA_DESCRICAO_VISUAL = "descricao_visual_ia"
COLUNA_ACAO = "acao_sugerida_ia"
COLUNA_ERRO = "erro_analise_ia"


def carregar_env(caminho: Path) -> None:
    if not caminho.exists():
        return

    for linha in caminho.read_text(encoding="utf-8").splitlines():
        linha = linha.strip()
        if not linha or linha.startswith("#") or "=" not in linha:
            continue

        chave, valor = linha.split("=", 1)
        chave = chave.strip()
        valor = valor.strip().strip('"').strip("'")
        if chave and chave not in os.environ:
            os.environ[chave] = valor


def carregar_env_automatico() -> None:
    pasta_script = Path(__file__).resolve().parent
    carregar_env(pasta_script.parent / ".env")
    carregar_env(pasta_script / ".env")
    carregar_env(Path.cwd() / ".env")


def ler_csv(caminho: Path) -> List[Dict[str, str]]:
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(4096)
        arquivo.seek(0)
        try:
            dialecto = csv.Sniffer().sniff(amostra, delimiters=",;") if amostra else csv.excel
        except csv.Error:
            dialecto = csv.excel
        leitor = csv.DictReader(arquivo, dialect=dialecto)
        if not leitor.fieldnames:
            raise ValueError("CSV de entrada esta vazio ou sem cabecalho")
        return [dict(linha) for linha in leitor]


def valor(linha: Dict[str, str], *nomes: str) -> str:
    for nome in nomes:
        if nome in linha and linha[nome]:
            return str(linha[nome]).strip()
    return ""


def baixar_imagem(image_url: str) -> Path:
    suffix = Path(image_url.split("?")[0]).suffix or ".jpg"
    destino = Path(tempfile.gettempdir()) / f"agente_ia_{abs(hash(image_url))}{suffix}"
    if not destino.exists():
        urllib.request.urlretrieve(image_url, destino)
    return destino


def imagem_para_data_url(caminho: Path) -> str:
    mime_type = mimetypes.guess_type(caminho.name)[0] or "image/jpeg"
    conteudo = base64.b64encode(caminho.read_bytes()).decode("utf-8")
    return f"data:{mime_type};base64,{conteudo}"


def obter_imagem_para_ia(linha: Dict[str, str], pasta_base: Path, baixar_links: bool) -> Optional[str]:
    image_path = valor(linha, "image_path", "caminho_imagem", "foto_path")
    image_url = valor(linha, "image_url", "url_imagem", "foto_url", "link_foto")

    if image_path:
        caminho = Path(image_path)
        if not caminho.is_absolute():
            caminho = pasta_base / caminho
        if not caminho.exists():
            raise FileNotFoundError(f"imagem local nao encontrada: {caminho}")
        return imagem_para_data_url(caminho)

    if image_url:
        if baixar_links:
            caminho = baixar_imagem(image_url)
            return imagem_para_data_url(caminho)
        return image_url

    return None


def extrair_json(texto: str) -> Dict[str, object]:
    texto = texto.strip()
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        inicio = texto.find("{")
        fim = texto.rfind("}")
        if inicio == -1 or fim == -1 or fim <= inicio:
            raise
        return json.loads(texto[inicio : fim + 1])


def imagem_data_url_para_base64(imagem: str) -> str:
    if not imagem.startswith("data:"):
        raise ValueError("Ollama precisa da imagem baixada em base64. Rode sem --nao-baixar-links.")
    return imagem.split(",", 1)[1]


def chamar_openai(prompt: str, imagem: str) -> str:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise ValueError("OPENAI_API_KEY nao configurada no .env")

    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    modelo = os.getenv("OPENAI_MODEL", "gpt-4.1")
    resposta = client.responses.create(
        model=modelo,
        input=[
            {
                "role": "user",
                "content": [
                    {"type": "input_text", "text": prompt},
                    {"type": "input_image", "image_url": imagem, "detail": "low"},
                ],
            }
        ],
    )
    return resposta.output_text


def chamar_ollama(prompt: str, imagem: str) -> str:
    from ollama import chat

    modelo = os.getenv("OLLAMA_MODEL", "gemma4:31b-cloud")
    response = chat(
        model=modelo,
        messages=[
            {
                "role": "user",
                "content": prompt,
                "images": [imagem_data_url_para_base64(imagem)],
            }
        ],
    )

    if hasattr(response, "message") and hasattr(response.message, "content"):
        return response.message.content
    return response["message"]["content"]


def chamar_ia(prompt: str, imagem: str, provedor_ia: str) -> str:
    if provedor_ia == "ollama":
        return chamar_ollama(prompt, imagem)
    if provedor_ia == "openai":
        return chamar_openai(prompt, imagem)
    raise ValueError(f"provedor de IA invalido: {provedor_ia}")


def montar_contexto(linha: Dict[str, str]) -> str:
    campos_interessantes = [
        "atividade_id",
        "tarefa_id",
        "colaborador",
        "data",
        "tarefa_nome",
        "checklist_nome",
        "checklist_descricao",
        "justificativa",
        "grupo",
        "motivo",
        "image_url",
    ]

    partes = []
    for campo in campos_interessantes:
        conteudo = valor(linha, campo)
        if conteudo:
            partes.append(f"{campo}: {conteudo}")
    return "\n".join(partes) if partes else "Sem contexto textual estruturado."


def analisar_linha_com_ia(linha: Dict[str, str], pasta_base: Path, baixar_links: bool, provedor_ia: str) -> Dict[str, str]:
    imagem = obter_imagem_para_ia(linha, pasta_base, baixar_links)
    justificativa = valor(linha, "justificativa")

    if not imagem:
        return {
            COLUNA_ANALISADA: "sem imagem",
            COLUNA_GRUPO: "sem_imagem",
            COLUNA_CONFIANCA: "1.00",
            COLUNA_MOTIVO: "Nao existe image_url ou image_path para a IA validar.",
            COLUNA_DESCRICAO_VISUAL: "",
            COLUNA_ACAO: "recusar por falta de comprovacao",
            COLUNA_ERRO: "",
        }

    contexto = montar_contexto(linha)

    prompt = f"""
Voce e um agente auditor de rondas operacionais.

Sua tarefa e comparar:
1. a justificativa textual do colaborador;
2. o contexto da tarefa/ronda;
3. a imagem enviada como evidencia.

Decida se a evidencia visual parece coerente com a justificativa.

Use:
- aprovada: quando a imagem parece util e coerente com a justificativa.
- reprovada: quando a imagem e preta, vazia, inutil, sem relacao aparente, ou contradiz a justificativa.
- duvidosa: quando a imagem tem alguma informacao, mas nao da para confirmar bem.
- sem_imagem: quando nao houver imagem, mas esse caso normalmente ja sera tratado antes.

Contexto da linha:
{contexto}

Justificativa principal:
{justificativa or "nao informada"}

Responda somente em JSON valido neste formato:
{{
  "analisada_por_ia": "aprovada|reprovada|duvidosa|sem_imagem",
  "confianca": 0.0,
  "motivo": "explicacao curta",
  "descricao_visual": "o que a imagem parece mostrar",
  "acao_sugerida": "aceitar|revisar|recusar"
}}
""".strip()

    resposta_texto = chamar_ia(prompt, imagem, provedor_ia)
    dados = extrair_json(resposta_texto)
    analisada = str(dados.get("analisada_por_ia", "duvidosa")).strip().lower()
    if analisada not in {"aprovada", "reprovada", "duvidosa", "sem_imagem"}:
        analisada = "duvidosa"

    confianca = float(dados.get("confianca", 0.0))
    confianca = max(0.0, min(1.0, confianca))

    return {
        COLUNA_ANALISADA: analisada,
        COLUNA_GRUPO: analisada,
        COLUNA_CONFIANCA: f"{confianca:.2f}",
        COLUNA_MOTIVO: str(dados.get("motivo", "")).strip(),
        COLUNA_DESCRICAO_VISUAL: str(dados.get("descricao_visual", "")).strip(),
        COLUNA_ACAO: str(dados.get("acao_sugerida", "")).strip(),
        COLUNA_ERRO: "",
    }


def salvar_xlsx(linhas: List[Dict[str, str]], saida: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    saida.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Analise IA"

    colunas = list(linhas[0].keys()) if linhas else []
    sheet.append(colunas)

    for linha in linhas:
        sheet.append([linha.get(coluna, "") for coluna in colunas])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    cores = {
        "aprovada": "DCFCE7",
        "reprovada": "FEE2E2",
        "duvidosa": "FEF3C7",
        "sem imagem": "E5E7EB",
        "sem_imagem": "E5E7EB",
    }

    if COLUNA_ANALISADA in colunas:
        indice = colunas.index(COLUNA_ANALISADA) + 1
        for row in range(2, sheet.max_row + 1):
            status = str(sheet.cell(row=row, column=indice).value or "").lower()
            fill = PatternFill("solid", fgColor=cores.get(status, "FFFFFF"))
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).fill = fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for coluna in range(1, sheet.max_column + 1):
        letra = get_column_letter(coluna)
        maior = 12
        for cell in sheet[letra]:
            maior = max(maior, len(str(cell.value or "")) + 2)
        sheet.column_dimensions[letra].width = min(maior, 55)

    workbook.save(saida)


def analisar_csv(entrada: Path, saida: Path, limite: Optional[int], baixar_links: bool, provedor_ia: str) -> None:
    linhas = ler_csv(entrada)
    pasta_base = entrada.parent
    linhas_saida = []

    for indice, linha in enumerate(linhas, start=1):
        if limite and indice > limite:
            linha_saida = {**linha}
            linha_saida[COLUNA_ANALISADA] = "nao analisada"
            linha_saida[COLUNA_GRUPO] = ""
            linha_saida[COLUNA_CONFIANCA] = ""
            linha_saida[COLUNA_MOTIVO] = "Limite de analise atingido."
            linha_saida[COLUNA_DESCRICAO_VISUAL] = ""
            linha_saida[COLUNA_ACAO] = ""
            linha_saida[COLUNA_ERRO] = ""
            linhas_saida.append(linha_saida)
            continue

        try:
            analise = analisar_linha_com_ia(linha, pasta_base, baixar_links, provedor_ia)
        except Exception as exc:
            analise = {
                COLUNA_ANALISADA: "erro",
                COLUNA_GRUPO: "erro",
                COLUNA_CONFIANCA: "0.00",
                COLUNA_MOTIVO: "A IA nao conseguiu analisar esta linha.",
                COLUNA_DESCRICAO_VISUAL: "",
                COLUNA_ACAO: "revisar",
                COLUNA_ERRO: str(exc),
            }

        linhas_saida.append({**linha, **analise})
        print(f"Analisado {indice}/{len(linhas)}: {analise[COLUNA_ANALISADA]}")

    salvar_xlsx(linhas_saida, saida)
    print(f"Arquivo gerado: {saida}")


def criar_parser() -> argparse.ArgumentParser:
    carregar_env_automatico()

    parser = argparse.ArgumentParser(description="Agente de IA para validar justificativa + foto de rondas.")
    parser.add_argument("--entrada", required=True, help="CSV gerado pelo auditor ou exportado do BI")
    parser.add_argument("--saida", required=True, help="Arquivo XLSX final com a coluna 'analisada por IA'")
    parser.add_argument("--limite", type=int, help="Limita a quantidade de linhas analisadas, util para teste")
    parser.add_argument(
        "--nao-baixar-links",
        action="store_true",
        help="Envia image_url direto para a IA em vez de baixar e enviar como base64",
    )
    parser.add_argument(
        "--provedor-ia",
        choices=["openai", "ollama"],
        default=(os.getenv("IA_PROVIDER") or os.getenv("AI_PROVIDER") or "openai").lower(),
        help="Provedor usado para analisar justificativa + imagem",
    )
    return parser


def main() -> None:
    args = criar_parser().parse_args()
    analisar_csv(
        entrada=Path(args.entrada),
        saida=Path(args.saida),
        limite=args.limite,
        baixar_links=not args.nao_baixar_links,
        provedor_ia=args.provedor_ia,
    )


if __name__ == "__main__":
    main()
